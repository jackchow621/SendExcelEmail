# -*- coding:utf-8 -*-
import time

import openpyxl

from config import SHEET_TITLE


def get_excel(data, field, file):
    # create a workbook
    new = openpyxl.Workbook()
    # create a sheet
    sheet = new.active
    # named the sheet
    sheet.title = SHEET_TITLE
    # write the field into the first row of sheet
    for col in range(len(field)):
        _ = sheet.cell(row=1, column=col + 1, value=u'%s' % field[col][0])
    #  write the data into the sheet
    for row in range(len(data)):
        for col in range(len(field)):
            if 'time' in field[col][0] or 'date' in field[col][0]:
                time_local = time.localtime(data[row][col])
                data[row][col] = time.strftime("%Y-%m-%d %H:%M:%S", time_local)
                # print data[row][col]
            _ = sheet.cell(row=row + 2, column=col + 1, value=u'%s' % data[row][col])
    # save the workbook
    newworkbook = new.save(file)
    return newworkbook
